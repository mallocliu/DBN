from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QMessageBox, QMenu
from DBN import runDBN
from mainwindow import Ui_MainWindow
from mythread import TrainThread, TestThread
from openpyxl import Workbook, load_workbook
import os
translate_result = {
    1: '优秀',
    2: '良好',
    3: '中等',
    4: '合格',
    5: '不合格'
}

result_int = {
    '优秀': 1,
    '良好': 2,
    '中等': 3,
    '合格': 4,
    '不合格': 5,
}


class MainWindow(Ui_MainWindow):

    def __init__(self, window):
        super().setupUi(window)
        self.dbn = []
        self.trainFileName = ''
        self.evaluateFileName = ''
        self.trThread = None
        self.teThread = []
        self.numIndexEvaluated = 0
        self.lineEdit_batch.setText('4')
        self.lineEdit_lr.setText('0.01')
        self.lineEdit_epoch.setText('100')
        self.tree_index.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked)
        self.tree_index.header().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        self.lineEdit_indexName.setReadOnly(True)
        # signal-slot connection

        self.tree_index.customContextMenuRequested.connect(self.createContextMenu)
        self.push_confirmindex.clicked.connect(self.confirmindex)
        self.push_selectIndex.clicked.connect(self.loadindex)
        self.push_saveIndex.clicked.connect(self.saveIndex)
        self.push_saveResult.clicked.connect(self.saveResult)
        self.push_saveCurrentindex.clicked.connect(self.saveCurrentindex)
        self.push_clearCurrentindex.clicked.connect(self.clearCurrentindex)
        self.push_checkIndex.clicked.connect(self.handle_checkIndex)
        self.groupBox_dbn.clicked.connect(self.handle_dbn_checked)
        self.groupBox_weight.clicked.connect(self.handle_weight_checked)
        self.push_saveWeight.clicked.connect(self.handle_saveWeight)

        self.pushButton_selectTrain.clicked.connect(self.handle_selectTrain)
        self.push_selectEvaluate.clicked.connect(self.handle_selectEvaluate)

        self.command_startTrain.clicked.connect(self.handle_startTrain)
        self.command_startEvaluate.clicked.connect(self.handle_startEvaluate)

        self.pushButton_saveModel.clicked.connect(self.handle_saveModel)
        self.pushButton_loadModel.clicked.connect(self.handle_loadModel)

    def createContextMenu(self):
        '''''
        创建右键菜单
        '''
        self.contextMenu = QMenu()
        if not self.tree_index.topLevelItemCount():
            self.action_newrootindex = self.contextMenu.addAction("新建根指标")
            self.action_newrootindex.triggered.connect(self.actionHandler_newrootindex)
            self.contextMenu.exec_(QtGui.QCursor.pos())
        else:
            if self.tree_index.currentItem() != self.tree_index.topLevelItem(0):
                self.action_newsamelevelindex = self.contextMenu.addAction("新建同级指标")
                self.action_newsamelevelindex.triggered.connect(self.actionHandler_newsamelevelindex)

            self.action_modifyindex = self.contextMenu.addAction('修改指标')
            self.action_newsublevelindex = self.contextMenu.addAction('新建子指标')
            self.action_deleteindex = self.contextMenu.addAction('删除该指标')
            self.action_modifyindex.triggered.connect(self.actionHandler_modifyindex)
            self.action_newsublevelindex.triggered.connect(self.actionHandler_newsublevelindex)
            self.action_deleteindex.triggered.connect(self.actionHandler_deleteindex)
            self.contextMenu.exec_(QtGui.QCursor.pos())

    def actionHandler_newsamelevelindex(self):
        from PyQt5.QtCore import Qt
        newwidget = QtWidgets.QTreeWidgetItem()
        newwidget.setFlags(Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable)
        parentwidget = self.tree_index.currentItem().parent()
        parentwidget.addChild(newwidget)

    def actionHandler_newsublevelindex(self):
        from PyQt5.QtCore import Qt
        newwidget = QtWidgets.QTreeWidgetItem()
        newwidget.setFlags(Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable)
        curitem = self.tree_index.currentItem()
        curitem.addChild(newwidget)
        curitem.setExpanded(True)

    def actionHandler_modifyindex(self):
        from PyQt5.QtCore import Qt
        curWidget = self.tree_index.currentItem()
        curWidget.setFlags(Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable)

    def actionHandler_deleteindex(self):
        import sip
        curitem = self.tree_index.currentItem()
        sip.delete(curitem)

    def actionHandler_newrootindex(self):
        from PyQt5.QtCore import Qt
        rootwidget = QtWidgets.QTreeWidgetItem()
        rootwidget.setFlags(Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable)
        self.tree_index.addTopLevelItem(rootwidget)

    def addSubindex(self, curWidget):
        subIndexNum = curWidget.childCount()
        subIndexWidgets = [curWidget.child(c) for c in range(subIndexNum)]
        self.combo_subIndex.clear()
        for widget in subIndexWidgets:
            self.combo_subIndex.addItem(widget.text(0))

    def handle_saveModel(self):
        import pickle, os
        opendialog = QtWidgets.QFileDialog()
        name = opendialog.getSaveFileName(None, 'Save Model...',
                                          os.getcwd(),
                                          'Model Files (*.pkl)')[0]
        with open(name, 'wb') as f:
            pickle.dump(self.dbn[-1], f)

    def handle_loadModel(self):
        import pickle, os
        opendialog = QtWidgets.QFileDialog()
        name = opendialog.getOpenFileName(None, 'Load Model...',
                                          os.getcwd(),
                                          'Model Files (*.pkl)')[0]
        with open(name, 'rb') as f:
            dbn = pickle.load(f)
            if dbn.rootindex != self.tree_index.currentItem().text(0):
                QMessageBox.warning(QMessageBox(), 'loading error', '加载的模型名称与指标名称不一致')
            else:
                self.setParameter(dbn)
                names = [node.rootindex for node in self.dbn]
                if name not in names:
                    self.dbn.append(dbn)
                else:
                    pos = names.index(name)
                    self.dbn[pos] = dbn

    def setParameter(self, dbn: runDBN):
        self.lineEdit_lr.setText(str(dbn.pretrain_lr))
        self.lineEdit_epoch.setText(str(dbn.pretrain_epoch))
        self.lineEdit_batch.setText(str(dbn.batch_size))
        l = len(dbn.hidden_layer_sizes)
        if l>=1:
            self.checkBox_lv1.setChecked(True)
            self.lineEdit_lv1.setText(str(dbn.hidden_layer_sizes[0]))
        if l>=2:
            self.checkBox_lv2.setChecked(True)
            self.lineEdit_lv2.setText(str(dbn.hidden_layer_sizes[1]))
        if l>=3:
            self.checkBox_lv3.setChecked(True)
            self.lineEdit_lv3.setText(str(dbn.hidden_layer_sizes[2]))

    def handle_weight_checked(self):
        self.groupBox_dbn.setChecked(False)
        self.addSubindex(self.tree_index.currentItem())

    def handle_dbn_checked(self):
        self.groupBox_weight.setChecked(False)

    def handle_checkIndex(self):
        treeiter = QtWidgets.QTreeWidgetItemIterator(self.tree_index)
        flag = True
        while treeiter.value():
            curWidget = treeiter.value()
            if curWidget.text(2) == '定性':
                cnt = curWidget.childCount()
                if cnt:
                    if curWidget.text(1) == '' and curWidget != self.tree_index.topLevelItem(0):
                        QMessageBox.warning(QMessageBox(), 'error',
                                            '指标{}未指定权值'.format(curWidget.text(0))
                                            )
                        flag = False
                    sum = 0.0
                    for i in range(cnt):
                        if curWidget.child(i).text(1) == '': continue
                        if curWidget.text(3) == '否' and curWidget.child(i).text(2) == '定性':
                            QMessageBox.warning(QMessageBox(), 'error',
                                                '指标{}为dbn计算得出，子指标必须为定量指标'.format(curWidget.text(0))
                                                )
                            flag = False
                        sum += float(curWidget.child(i).text(1))
                    if sum != 1 and sum != 0:
                        QMessageBox.warning(QMessageBox(), 'error',
                                            '指标{}子指标权值之和不等于1'.format(curWidget.text(0))
                                            )
                        flag = False
                else:
                    if curWidget.text(3):
                        QMessageBox.warning(QMessageBox(), 'error',
                                            '指标{}无需计算'.format(curWidget.text(0))
                                            )
                        flag = False
            else:
                cnt = curWidget.childCount()
                if cnt:
                    QMessageBox.warning(QMessageBox(), 'error',
                                        '指标{}为定量指标，不可能包含子指标'.format(curWidget.text(0))
                                        )
                    flag = False
            treeiter += 1
        if flag:
            QMessageBox.information(QMessageBox(), 'correct', '校验通过')

    def handle_selectTrain(self):
        opendialog = QtWidgets.QFileDialog()
        self.trainFileName = opendialog.getOpenFileName(None, 'Load Index Sample File...',
                                          os.getcwd(),
                                          '(*.xlsx)')[0]

    def handle_selectEvaluate(self):
        opendialog = QtWidgets.QFileDialog()
        self.evaluateFileName = opendialog.getOpenFileName(None, 'Load Index Evaluate File...',
                                                   os.getcwd(),
                                                   '(*.xlsx)')[0]


    def handle_startTrain(self):
        newDBN = runDBN()
        newDBN.trainFilePath, newDBN.rootindex, newDBN.indexname, newDBN.hidden_layer_sizes, \
        newDBN.batch_size, newDBN.pretrain_epoch, newDBN.pretrain_lr, = self.get_train_parameters()
        self.dbn.append(newDBN)
        self.trThread = TrainThread(self.dbn[-1])
        self.trThread.start()

    def cal_indexEvaluateNum(self):
        treeiter = QtWidgets.QTreeWidgetItemIterator(self.tree_index)
        self.numIndexEvaluated = 0
        while treeiter.value():
            curwidget = treeiter.value()
            if curwidget.text(3)=='否' or (curwidget.text(2)=='定性' and curwidget.text(3) == ''):
                self.numIndexEvaluated +=1
            treeiter += 1

    def handle_startEvaluate(self):
        ws = load_workbook(filename=self.evaluateFileName, read_only=True)['Sheet1']
        if not self.isValidEvaluateFile(ws):
            return
        column = 'B'
        header = []
        self.cal_indexEvaluateNum()
        self.table_result.setRowCount(ws.max_row - 6)
        self.table_result.setColumnCount(self.numIndexEvaluated)
        tmp = 0
        while True:
            indexname = ws['{}5'.format(column)].value
            indextype = ws['{}6'.format(column)].value
            if indexname is None:
                break
            if indextype == '定量数据':
                cnt = self.evaluate_DBN(indexname, tmp, ws, column, header)
                column = chr(ord(column) + cnt)
                tmp += 1
            else:
                self.evaluate_weight(indexname, tmp, ws, column, header)
                column = chr(ord(column) + 1)
                tmp += 1

    def handle_evaluate_rootIndex(self, header):
        sum = self.calc_result(self.tree_index.topLevelItem(0))
        header.append(self.tree_index.topLevelItem(0).text(0))
        self.table_result.setColumnCount(self.table_result.columnCount() + 1)
        self.table_result.setHorizontalHeaderLabels(header)
        for i in range(len(sum)):
            newwidget = QtWidgets.QTableWidgetItem(translate_result[round(sum[i])])
            newwidget.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            self.table_result.setItem(i, self.table_result.columnCount() - 1, newwidget)
        QMessageBox.information(QMessageBox(), '', '评估完成！')

    def saveResult(self):
        opendialog = QtWidgets.QFileDialog()
        name = opendialog.getSaveFileName(None, 'Save Evaluate Result...',
                                          os.getcwd(),
                                          '(*.xlsx)')[0]
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        max_row = self.table_result.rowCount()
        max_col = self.table_result.columnCount()
        header = ['试验序号']
        for i in range(max_col):
            header.append(self.table_result.horizontalHeaderItem(i).text())
        ws.append(header)
        for r in range(max_row):
            tmp = [self.table_result.item(r, c).text() for c in range(max_col)]
            ws.append([str(r)] + tmp)
        wb.save(name)


    def findcol(self, name):
        max_col = self.table_result.columnCount()
        for i in range(max_col):
            if self.table_result.horizontalHeaderItem(i).text() == name:
                return i

    def calc_result(self, node):
        ans = [0 for _ in range(self.table_result.rowCount())]
        if (node.text(3) == '否') or (node.text(2) == '定性' and node.text(3) == ''):
            col = self.findcol(node.text(0))
            tmp = [result_int[self.table_result.item(r, col).text()] * float(node.text(1))
                   for r in range(self.table_result.rowCount())]
            return tmp
        else:
            cnt = node.childCount()
            for i in range(cnt):
                tmp = self.calc_result(node.child(i))
                ans = [sum(x) for x in zip(ans, tmp)]
        return ans

    def evaluate_DBN(self, indexname, pos, ws, column, header):
        curwidget = self.tree_index.findItems(indexname,
                                              QtCore.Qt.MatchFixedString |
                                              QtCore.Qt.MatchRecursive)[0].parent()
        curwidget_index = curwidget.text(0)
        names = [node.rootindex for node in self.dbn]
        index = names.index(curwidget_index)
        curdbn = self.dbn[index]  # find the dbn we shuold use
        header.append(curwidget_index)
        self.table_result.setHorizontalHeaderLabels(header)
        newThread = TestThread(curdbn, pos, ws, column, header)
        newThread.evaluate_dbn_finished.connect(self.handle_one_evaluate_dbn_finished)
        self.teThread.append(newThread)
        self.teThread[-1].start()
        return curwidget.childCount()

    def handle_one_evaluate_dbn_finished(self,  label, pos, header):
        maxrow = self.table_result .rowCount()
        label = label[:maxrow]
        for i in range(maxrow):
            newWidget = QtWidgets.QTableWidgetItem(translate_result[label[i] + 1])
            newWidget.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            self.table_result.setItem(i, pos, newWidget)
        nullcells = self.table_result.findItems('',
                                    QtCore.Qt.MatchRecursive | QtCore.Qt.MatchFixedString)
        if not nullcells:
            self.handle_evaluate_rootIndex(header)

    def evaluate_weight(self, indexname, pos, ws, column, header):
        header.append(indexname)
        self.table_result.setHorizontalHeaderLabels(header)

        for i in range(7, ws.max_row + 1):
            val = ws['{}{}'.format(column, i)].value
            curcell = QtWidgets.QTableWidgetItem(translate_result[val])
            curcell.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
            self.table_result.setItem(i - 7, pos, curcell)

    def isValidEvaluateFile(self, ws):
        flag = True
        if ws['B2'].value != self.tree_index.topLevelItem(0).text(0):
            QMessageBox.warning(QMessageBox(), 'error', '待评估指标与文件中不一致')
            flag = False
        if ws['D2'].value != '评估数据':
            QMessageBox.warning(QMessageBox(), 'error', '该文件不是评估文件')
            flag = False
        return flag

    def get_train_parameters(self):
        name = self.trainFileName
        curWidget = self.tree_index.currentItem()
        indexname = set()
        rootindex = curWidget.text(0)
        for i in range(curWidget.childCount()):
            indexname.add(curWidget.child(i).text(0))
        batch_size = int(self.lineEdit_batch.text())
        pretrain_epoch = int(self.lineEdit_epoch.text())
        pretrain_lr = float(self.lineEdit_lr.text())
        hidden_layer_sizes = []
        if self.checkBox_lv1.isChecked():
            hidden_layer_sizes.append(int(self.lineEdit_lv1.text()))

        if self.checkBox_lv2.isChecked():
            hidden_layer_sizes.append(int(self.lineEdit_lv2.text()))

        if self.checkBox_lv3.isChecked():
            hidden_layer_sizes.append(int(self.lineEdit_lv3.text()))

        return name, rootindex, indexname, hidden_layer_sizes, \
               batch_size, pretrain_epoch, pretrain_lr

    def handle_saveWeight(self):
        indexId = self.combo_subIndex.currentIndex()
        curWidget = self.tree_index.currentItem().child(indexId)
        curWidget.setText(1, self.lineEdit_weight.text())

    def confirmindex(self):
        from PyQt5.QtCore import Qt
        treeiter = QtWidgets.QTreeWidgetItemIterator(self.tree_index)
        while treeiter.value():
            treeiter.value().setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsUserCheckable)
            treeiter += 1

        curWidget = self.tree_index.currentItem()

        self.tree_index.currentItemChanged.connect(
            lambda: self.lineEdit_indexName.setText(self.tree_index.currentItem().text(0)))
        self.tree_index.currentItemChanged.connect(
            lambda: self.produce_safe_option(self.tree_index.currentItem()))
        self.lineEdit_indexName.setText(curWidget.text(0))
        self.produce_safe_option(curWidget)

    def produce_safe_option(self, curWidget):
        rootWidget = self.tree_index.topLevelItem(0)
        self.combo_indexType.clear()
        if curWidget.text(2):
            self.combo_indexType.addItem(curWidget.text(2))
            if curWidget.text(3) == '是':
                self.groupBox_weight.setChecked(True)
                self.handle_weight_checked()
            elif curWidget.text(3) == '否':
                self.groupBox_dbn.setChecked(True)
                self.handle_dbn_checked()
            else:
                self.groupBox_dbn.setChecked(False)
                self.groupBox_dbn.setChecked(False)
        else:
            if curWidget == rootWidget:
                # root index can only be weight calculated
                self.combo_indexType.addItem('定性')
                self.groupBox_weight.setChecked(True)
                self.handle_weight_checked()
            else:
                self.combo_subIndex.clear()
                self.combo_indexType.addItem('定性')
                self.combo_indexType.addItem('定量')
                self.groupBox_weight.setChecked(False)
                self.groupBox_dbn.setChecked(False)

    def loadindex(self):
        import os
        opendialog = QtWidgets.QFileDialog()
        name = opendialog.getOpenFileName(None, 'Load Index File...',
                                          os.getcwd(),
                                          'index Files (*.index)')[0]
        indexFile = QtCore.QFile(name)
        indexFile.open(QtCore.QIODevice.ReadOnly)
        data = QtCore.QDataStream(indexFile)
        child = QtWidgets.QTreeWidgetItem(self.tree_index.invisibleRootItem())
        child.read(data)
        num_childs = data.readUInt32()
        self.restore_item(data, child, num_childs)

    def restore_item(self, data, item, num_childs):
        for i in range(num_childs):
            child = QtWidgets.QTreeWidgetItem(item)
            child.read(data)
            cnt = data.readUInt32()
            self.restore_item(data, child, cnt)

    def saveIndex(self):
        import os
        opendialog = QtWidgets.QFileDialog()
        name = opendialog.getSaveFileName(None, 'Save File...',
                                          os.getcwd(),
                                          'index Files (*.index)')[0]
        indexFile = QtCore.QFile(name)
        indexFile.open(QtCore.QIODevice.WriteOnly)
        data = QtCore.QDataStream(indexFile)
        self.save_item(self.tree_index.invisibleRootItem(), data)

    def saveCurrentindex(self):
        curWidget = self.tree_index.currentItem()
        curWidget.setText(2, self.combo_indexType.currentText())
        if self.groupBox_weight.isChecked():
            curWidget.setText(3, '是')
        elif self.groupBox_dbn.isChecked():
            curWidget.setText(3, '否')
        else:
            curWidget.setText(3, '')

    def clearCurrentindex(self):
        curWidget = self.tree_index.currentItem()
        curWidget.setText(2, '')
        curWidget.setText(3, '')
        if self.groupBox_weight.isChecked():
            cnt = curWidget.childCount()
            for i in range(cnt):
                curWidget.child(i).setText(1, '')

    def save_item(self, item, data):
        cnt = item.childCount()
        for i in range(cnt):
            child = item.child(i)
            child.write(data)
            data.writeUInt32(child.childCount())
            self.save_item(child, data)
