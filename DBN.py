"""
"""
from __future__ import print_function, division

import sys
import timeit

import numpy
import theano
import theano.tensor as T
from PyQt5.QtCore import QObject, pyqtSignal
from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook
from theano.sandbox.rng_mrg import MRG_RandomStreams

from logistic_sgd import LogisticRegression
from mlp import HiddenLayer
from rbm import RBM


# start-snippet-1


class DBN(object):
    """Deep Belief Network

    A deep belief network is obtained by stacking several RBMs on top of each
    other. The hidden layer of the RBM at layer `i` becomes the input of the
    RBM at layer `i+1`. The first layer RBM gets as input the input of the
    network, and the hidden layer of the last RBM represents the output. When
    used for classification, the DBN is treated as a MLP, by adding a logistic
    regression layer on top.
    """

    def __init__(self, numpy_rng, theano_rng=None, n_ins=3,
                 hidden_layers_sizes=[6, 12], n_outs=5):
        """This class is made to support a variable number of layers.

        :type numpy_rng: numpy.random.RandomState
        :param numpy_rng: numpy random number generator used to draw initial
                    weights

        :type theano_rng: theano.tensor.shared_randomstreams.RandomStreams
        :param theano_rng: Theano random generator; if None is given one is
                           generated based on a seed drawn from `rng`

        :type n_ins: int
        :param n_ins: dimension of the input to the DBN

        :type hidden_layers_sizes: list of ints
        :param hidden_layers_sizes: intermediate layers size, must contain
                               at least one value

        :type n_outs: int
        :param n_outs: dimension of the output of the network
        """

        self.sigmoid_layers = []
        self.rbm_layers = []
        self.params = []
        self.n_layers = len(hidden_layers_sizes)

        assert self.n_layers > 0

        if not theano_rng:
            theano_rng = MRG_RandomStreams(numpy_rng.randint(2 ** 10))

        # allocate symbolic variables for the data

        # the data is presented as rasterized images
        self.x = T.matrix('x')

        # the labels are presented as 1D vector of [int] labels
        self.y = T.ivector('y')
        # end-snippet-1
        # The DBN is an MLP, for which all weights of intermediate
        # layers are shared with a different RBM.  We will first
        # construct the DBN as a deep multilayer perceptron, and when
        # constructing each sigmoidal layer we also construct an RBM
        # that shares weights with that layer. During pretraining we
        # will train these RBMs (which will lead to chainging the
        # weights of the MLP as well) During finetuning we will finish
        # training the DBN by doing stochastic gradient descent on the
        # MLP.

        for i in range(self.n_layers):
            # construct the sigmoidal layer

            # the size of the input is either the number of hidden
            # units of the layer below or the input size if we are on
            # the first layer
            if i == 0:
                input_size = n_ins
            else:
                input_size = hidden_layers_sizes[i - 1]

            # the input to this layer is either the activation of the
            # hidden layer below or the input of the DBN if you are on
            # the first layer
            if i == 0:
                layer_input = self.x
            else:
                layer_input = self.sigmoid_layers[-1].output

            sigmoid_layer = HiddenLayer(rng=numpy_rng,
                                        input=layer_input,
                                        n_in=input_size,
                                        n_out=hidden_layers_sizes[i],
                                        activation=T.nnet.sigmoid)

            # add the layer to our list of layers
            self.sigmoid_layers.append(sigmoid_layer)

            # its arguably a philosophical question...  but we are
            # going to only declare that the parameters of the
            # sigmoid_layers are parameters of the DBN. The visible
            # biases in the RBM are parameters of those RBMs, but not
            # of the DBN.
            self.params.extend(sigmoid_layer.params)

            # Construct an RBM that shared weights with this layer
            rbm_layer = RBM(numpy_rng=numpy_rng,
                            theano_rng=theano_rng,
                            input=layer_input,
                            n_visible=input_size,
                            n_hidden=hidden_layers_sizes[i],
                            W=sigmoid_layer.W,
                            hbias=sigmoid_layer.b)
            self.rbm_layers.append(rbm_layer)

        # We now need to add a logistic layer on top of the MLP
        self.logLayer = LogisticRegression(
            input=self.sigmoid_layers[-1].output,
            n_in=hidden_layers_sizes[-1],
            n_out=n_outs)
        self.params.extend(self.logLayer.params)

        # compute the cost for second phase of training, defined as the
        # negative log likelihood of the logistic regression (output) layer
        self.finetune_cost = self.logLayer.negative_log_likelihood(self.y)

        # compute the gradients with respect to the model parameters
        # symbolic variable that points to the number of errors made on the
        # minibatch given by self.x and self.y
        self.errors = self.logLayer.errors(self.y)
        self.predict = self.logLayer.y_pred

    def pretraining_functions(self, train_set_x, batch_size, k):
        '''
        Generates a list of functions, for performing one step of
        gradient descent at a given layer. The function will require
        as input the minibatch index, and to train an RBM you just
        need to iterate, calling the corresponding function on all
        minibatch indexes.

        :type train_set_x: theano.tensor.TensorType
        :param train_set_x: Shared var. that contains all datapoints used
                            for training the RBM
        :type batch_size: int
        :param batch_size: size of a [mini]batch
        :param k: number of Gibbs steps to do in CD-k / PCD-k
        '''

        # index to a [mini]batch
        index = T.lscalar('index')  # index to a minibatch
        learning_rate = T.scalar('lr')  # learning rate to use

        # begining of a batch, given `index`
        batch_begin = index * batch_size
        # ending of a batch given `index`
        batch_end = batch_begin + batch_size

        pretrain_fns = []
        for rbm in self.rbm_layers:
            # get the cost and the updates list
            # using CD-k here (persisent=None) for training each RBM.
            # TODO: change cost function to reconstruction error
            cost, updates = rbm.get_cost_updates(learning_rate,
                                                 persistent=None, k=k)

            # compile the theano function
            fn = theano.function(
                inputs=[index, theano.In(learning_rate, value=0.1)],
                outputs=cost,
                updates=updates,
                givens={
                    self.x: train_set_x[batch_begin:batch_end]
                }
            )
            # append `fn` to the list of functions
            pretrain_fns.append(fn)

        return pretrain_fns

    def build_finetune_functions(self, datasets, batch_size, learning_rate):
        '''Generates a function `train` that implements one step of
        finetuning, a function `validate` that computes the error on a
        batch from the validation set, and a function `test` that
        computes the error on a batch from the testing set

        :type datasets: list of pairs of theano.tensor.TensorType
        :param datasets: It is a list that contain all the datasets;
                        the has to contain three pairs, `train`,
                        `valid`, `test` in this order, where each pair
                        is formed of two Theano variables, one for the
                        datapoints, the other for the labels
        :type batch_size: int
        :param batch_size: size of a minibatch
        :type learning_rate: float
        :param learning_rate: learning rate used during finetune stage
        '''

        (train_set_x, train_set_y) = datasets[0]
        (test_set_x, test_set_y) = datasets[1]

        # compute number of minibatches for training, validation and testing
        n_test_batches = test_set_x.get_value(borrow=True).shape[0]
        n_test_batches //= batch_size

        index = T.lscalar('index')  # index to a [mini]batch

        # compute the gradients with respect to the model parameters
        gparams = T.grad(self.finetune_cost, self.params)

        # compute list of fine-tuning updates
        updates = []
        for param, gparam in zip(self.params, gparams):
            updates.append((param, param - gparam * learning_rate))

        train_fn = theano.function(
            inputs=[index],
            outputs=self.finetune_cost,
            updates=updates,
            givens={
                self.x: train_set_x[
                        index * batch_size: (index + 1) * batch_size
                        ],
                self.y: train_set_y[
                        index * batch_size: (index + 1) * batch_size
                        ]
            }
        )

        test_score_i = theano.function(
            [index],
            self.errors,
            givens={
                self.x: test_set_x[
                        index * batch_size: (index + 1) * batch_size
                        ],
                self.y: test_set_y[
                        index * batch_size: (index + 1) * batch_size
                        ]
            }
        )

        test_label_i = theano.function(
            [index],
            self.predict,
            givens={
                self.x: test_set_x[
                        index * batch_size: (index + 1) * batch_size
                        ],
            }
        )

        # Create a function that scans the entire validation set
        # def valid_score():
        #   return [valid_score_i(i) for i in range(n_valid_batches)]

        # Create a function that scans the entire test set
        def test_score():
            return [test_score_i(i) for i in range(n_test_batches)]

        def test_label():
            return [test_label_i(i) for i in range(n_test_batches)]

        return train_fn, test_score, test_label


class runDBN(QObject):
    train_finished = pyqtSignal()
    test_finished = pyqtSignal()
    loading_train_error = pyqtSignal(str)
    loading_test_error = pyqtSignal(str)
    scoretable = {
        0: 95,
        1: 85,
        2: 75,
        3: 65,
        4: 55,
    }

    translate_result = {
        0: '优秀',
        1: '良好',
        2: '中等',
        3: '合格',
        4: '不合格'
    }

    def __init__(self):
        super().__init__()
        self.dbn = DBN(numpy_rng=numpy.random.RandomState(123))
        self.datasets = []
        self.minval = None
        self.maxval = None
        self.train_finished.connect(self.handle_train_finished)
        self.loading_train_error.connect(self.handle_loading_train_error)
        self.n_train_batches = 0
        self.hidden_layer_sizes = []
        self.pretrain_epoch = 0
        self.pretrain_lr = 0.0
        self.batch_size = 0
        self.trainFilePath = ''
        self.totalresult = ''
        self.score = 0
        self.indexname = set()
        self.rootindex = ''

    def __setstate__(self, state):
        super().__init__()
        self.__dict__.update(state)

    def handle_loading_train_error(self, message):
        QMessageBox.warning(QMessageBox(), 'error', message)

    def handle_train_finished(self):
        QMessageBox.information(QMessageBox(), "提示", "训练已完成")

    def handle_test_finished(self):
        QMessageBox.information(QMessageBox(), "提示", "试验已完成")

    def isValidTrainData(self, ws):
        flag = True
        if ws['B2'].value != self.rootindex:
            self.loading_train_error.emit('文件名称与待训练指标名称不符')
            flag = False
        if ws['D2'].value != '样本数据':
            self.loading_train_error.emit('该文件数据不是样本数据')
            flag = False
        feaNum = len(self.indexname)
        columns = [chr(ord('B') + i) for i in range(feaNum)]  # data range in file which we need to read
        for column in columns:
            name = ws['{}5'.format(column)].value
            indextype = ws['{}6'.format(column)].value
            if name not in self.indexname:
                self.loading_train_error.emit('子指标名称与指标体系不一致')
                flag = False
            if indextype != '定量数据':
                self.loading_train_error.emit('子指标类型出现非定量数据')
                flag = False
        return flag

    def load_traindata(self, dataset):
        ''' Loads the dataset

        :type dataset: string
        :param dataset: the path to the dataset
        '''

        ws = load_workbook(filename=dataset, read_only=True)['Sheet1']
        if not self.isValidTrainData(ws):
            return
        feaNum = len(self.indexname)
        columns = [chr(ord('B') + i) for i in range(feaNum)] # data range in file which we need to read

        data = list()
        labels = list()
        for row in range(7, ws.max_row + 1):
            tmp = list()
            for column in columns:  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                tmp.append(ws[cell_name].value)
            data.append(tmp)
            cell_name = "{}{}".format(chr(ord(columns[-1]) + 1), row)
            labels.append(ws[cell_name].value - 1)
        data = numpy.array(data)
        labels = numpy.array(labels).T

        dataSize = data.shape
        minval = data.min(axis=0)
        maxval = data.max(axis=0)

        ###normaolization
        for row in range(dataSize[0]):
            for col in range(dataSize[1]):
                data[row][col] = (data[row][col] - minval[col]) / (maxval[col] - minval[col])

        train_set = (data, labels)
        train_set_x, train_set_y = shared_dataset(train_set)
        return train_set_x, train_set_y, minval, maxval

    # todo: write loading test data process
    def load_testdata(self, ws, column):

        ''' Loads the dataset

        :type dataset: string
        :param dataset: the path to the dataset
        '''

        #############
        # LOAD DATA #
        #############
        # Load the dataset
        # train_set, test_set format: tuple(input, target)
        # input is a numpy.ndarray of 2 dimensions (a matrix)
        # where each row corresponds to an example. target is a
        # numpy.ndarray of 1 dimension (vector) that has the same length as
        # the number of rows in the input. It should give the target
        # to the example with the same index in the input.
        sampleNum = ws.max_row - 6
        res = self.batch_size - sampleNum % self.batch_size
        if res == self.batch_size:
            res = 0
        labels = [0 for _ in range(sampleNum + res)]  # give all the test label with 0 just for completeness
        columns = [chr(ord(column) + _) for _ in range(len(self.indexname))]
        testdata = [ws['{}{}'.format(c, r)].value for r in range(7, ws.max_row+1) for c in columns]
        remdata = [ws['{}{}'.format(c, ws.max_row)].value for c in columns]
        for _ in range(res):
            testdata.extend(remdata)
        testdata = numpy.array(testdata).reshape(sampleNum + res, len(self.indexname))
        labels = numpy.array(labels)
        dataSize = testdata.shape

        minval = testdata.min(axis=0)
        maxval = testdata.max(axis=0)

        ###normaolization
        for col in range(dataSize[1]):
            for row in range(dataSize[0]):
                testdata[row][col] = (testdata[row][col] - min(minval[col], self.minval[col])) / \
                                     (max(maxval[col], self.maxval[col]) - min(minval[col], self.minval[col]))

        test_set = (testdata, labels)

        test_set_x, test_set_y = shared_dataset(test_set)
        return test_set_x, test_set_y

    def pretrain_DBN(self):
        train_set_x, train_set_y, minval, maxval = self.load_traindata(self.trainFilePath)
        # when we haven't trained this index, we trained it, otherwise refresh the datasets.

        self.datasets = [(train_set_x, train_set_y)]
        self.minval = minval
        self.maxval = maxval
        self.n_train_batches = train_set_x.get_value(borrow=True).shape[0] // self.batch_size

        # numpy random generator
        numpy_rng = numpy.random.RandomState(123)
        # construct the Deep Belief Network
        self.dbn = DBN(numpy_rng=numpy_rng, n_ins=len(self.indexname),
                       hidden_layers_sizes=self.hidden_layer_sizes,
                       n_outs=5)
        # n_out=5, for we only have A+ A B C D.
        k = 1
        print('... getting the pretraining functions')
        pretraining_fns = self.dbn.pretraining_functions(train_set_x=train_set_x,
                                                         batch_size=self.batch_size,
                                                         k=k)
        print('... pre-training the model')
        # Pre-train layer-wise
        for i in range(self.dbn.n_layers):
            for epoch in range(self.pretrain_epoch):
                c = []
                for batch_index in range(self.n_train_batches):
                    c.append(pretraining_fns[i](index=batch_index,
                                                lr=self.pretrain_lr))

        self.train_finished.emit()

    def test_DBN(self):
        finetune_lr = 0.1
        training_epochs = 1000
        # get the training, validation and testing function for the model
        print('... getting the finetuning functions')
        train_fn, test_model, predict_model = self.dbn.build_finetune_functions(
            datasets=self.datasets,
            batch_size=self.batch_size,
            learning_rate=finetune_lr
        )

        print('... finetuning the model')
        epoch = 0
        while (epoch < training_epochs):
            epoch = epoch + 1
            for minibatch_index in range(self.n_train_batches):
                train_fn(minibatch_index)

        predict_label = predict_model()
        last_label = []
        for eachbatch in predict_label:
            last_label.extend(eachbatch.tolist())

        return last_label

def shared_dataset(data_xy, borrow=True):
    """ Function that loads the dataset into shared variables

    The reason we store our dataset in shared variables is to allow
    Theano to copy it into the GPU memory (when code is run on GPU).
    Since copying data into the GPU is slow, copying a minibatch everytime
    is needed (the default behaviour if the data is not in a shared
    variable) would lead to a large decrease in performance.
    """

    data_x, data_y = data_xy

    shared_x = theano.shared(numpy.asarray(data_x,
                                           dtype=theano.config.floatX),
                             borrow=borrow)
    shared_y = theano.shared(numpy.asarray(data_y,
                                           dtype=theano.config.floatX),
                             borrow=borrow)
    # When storing data on the GPU it has to be stored as floats
    # therefore we will store the labels as ``floatX`` as well
    # (``shared_y`` does exactly that). But during our computations
    # we need them as ints (we use labels as index, and if they are
    # floats it doesn't make sense) therefore instead of returning
    # ``shared_y`` we will have to cast it to int. This little hack
    # lets ous get around this issue
    return shared_x, T.cast(shared_y, 'int32')
