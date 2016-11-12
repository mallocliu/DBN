import numpy
from openpyxl import load_workbook
import theano
import theano.tensor as T


def shared_dataset(data_xy, borrow=True):
    """ Function that loads the dataset into shared variables

    The reason we store our dataset in shared variables is to allow
    Theano to copy it into the GPU memory (when code is run on GPU).
    Since copying data into the GPU is slow, copying a minibatch everytime
    is needed (the default behaviour if the data is not in a shared
    variable) would lead to a large decrease in performance.
    """

    data_x, data_y = data_xy

    shared_x = theano.shared(numpy.asarray(data_x,
                                           dtype=theano.config.floatX),
                             borrow=borrow)
    shared_y = theano.shared(numpy.asarray(data_y,
                                           dtype=theano.config.floatX),
                             borrow=borrow)
    # When storing data on the GPU it has to be stored as floats
    # therefore we will store the labels as ``floatX`` as well
    # (``shared_y`` does exactly that). But during our computations
    # we need them as ints (we use labels as index, and if they are
    # floats it doesn't make sense) therefore instead of returning
    # ``shared_y`` we will have to cast it to int. This little hack
    # lets ous get around this issue
    return shared_x, T.cast(shared_y, 'int32')


def load_traindata(dataset):
    ''' Loads the dataset

    :type dataset: string
    :param dataset: the path to the dataset
    '''

    wb = load_workbook(filename=dataset, read_only=True)
    ws = wb['Sheet1']

    columns = [chr(ord('A') + i) for i in range(ws.max_column - 1)]
    data = list()
    labels = list()
    for row in range(2, ws.max_row + 1):
        tmp = list()
        for column in columns:  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            tmp.append(ws[cell_name].value)
        data.append(tmp)
        cell_name = "{}{}".format(chr(ord(columns[-1]) + 1), row)
        labels.append(ws[cell_name].value - 1)
    data = numpy.array(data)
    labels = numpy.array(labels)
    labels = labels.T

    dataSize = data.shape
    # print(dataSize)
    minval = data.min(axis=0)
    maxval = data.max(axis=0)

    ###normaolization
    for row in range(0, dataSize[0]):
        for col in range(0, dataSize[1]):
            data[row][col] = (data[row][col] - minval[col]) / (maxval[col] - minval[col])
    train_set = (data, labels)
    train_set_x, train_set_y = shared_dataset(train_set)

    return train_set_x, train_set_y


def load_testdata(dataset):
    ''' Loads the dataset

    :type dataset: string
    :param dataset: the path to the dataset
    '''

    #############
    # LOAD DATA #
    #############
    # Load the dataset
    # train_set, test_set format: tuple(input, target)
    # input is a numpy.ndarray of 2 dimensions (a matrix)
    # where each row corresponds to an example. target is a
    # numpy.ndarray of 1 dimension (vector) that has the same length as
    # the number of rows in the input. It should give the target
    # to the example with the same index in the input.

    wb = load_workbook(filename=dataset, read_only=True)
    ws = wb['Sheet1']
    data = list()
    labels = [0]  # give all the test label with 0 just for completeness
    columns = [chr(ord('A') + i) for i in range(ws.max_column)]
    tmp = []
    for column in columns:
        cell_name = "{}{}".format(column, 2)  # give 2 because we only have 1 test sample
        tmp.append(ws[cell_name].value)
    data.append(tmp)
    data = numpy.array(data)
    labels = numpy.array(labels)
    dataSize = data.shape

    # print(dataSize)
    minval = data.min(axis=0)
    maxval = data.max(axis=0)

    ###normaolization
    for row in range(0, dataSize[0]):
        for col in range(0, dataSize[1]):
            data[row][col] = (data[row][col] - minval[col]) / (maxval[col] - minval[col])

    test_set = (data, labels)

    test_set_x, test_set_y = shared_dataset(test_set)
    return test_set_x, test_set_y
